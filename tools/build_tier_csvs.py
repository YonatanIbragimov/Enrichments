#!/usr/bin/env python3.11
"""
Build per-tier source CSVs from the master xlsx.

Reads source/fl_str_pm_companies_10plus_units.xlsx, applies the same cleaning
rules + 6-signal scoring as the in-browser JS, and writes:

    <tier>/source_cleaned_scored.csv

…ready for enrich_batch.py. Run once when the master xlsx changes; idempotent.

Usage:
    /opt/homebrew/bin/python3.11 tools/build_tier_csvs.py
"""

import csv
import re
import sys
from pathlib import Path

# Use whatever python3.11 has openpyxl — the browser-use venv has it.
sys.path.insert(0, "/Users/trinity/.browser-use-env/lib/python3.11/site-packages")
import openpyxl  # noqa: E402

ROOT   = Path(__file__).resolve().parent.parent
XLSX   = ROOT / "source" / "fl_str_pm_companies_10plus_units.xlsx"

OUT_COLS = [
    "Rank","Licensee Name","Total Units","Unique Addresses","Counties","Cities Served","Phone",
    "Score","Owner Name","Owner Email","Owner LinkedIn","Website","Notes",
]

# Map sheet name → (output dir, tier_min, tier_max-or-None)
TIERS = {
    "10-24 Listings":   ("10-24 Listings",  10,   24),
    "25-49 Listings":   ("25-49 Listings",  25,   49),
    "50-99 Listings":   ("50-99 Listings",  50,   99),
    "100-199 Listings": ("100-199 Listings",100,  199),
    "200+ Listings":    ("200+ Listings",   200,  None),
}

ASSOC_RE = re.compile(r"\b(ASSOCIATION|ASSOC|ASN|CONDOMINIUM|CONDO|HOA|HOMEOWNER|TIMESHARE)\b", re.I)
ENTITY_RE = re.compile(r"\b(LLC|INC|CORP|LTD|LP|L\.?L\.?C\.?)\b", re.I)
PMKW_RE   = re.compile(r"\b(MANAGEMENT|REALTY|PROPERTIES|RENTALS|VACATION|REAL ESTATE)\b", re.I)


def clean(row):
    name = (row.get("Licensee Name") or "").strip()
    if not name:                            return None, "Blank licensee name"
    if ASSOC_RE.search(name):               return None, "Association/HOA/Condo/Timeshare"
    try: uniq = int(row.get("Unique Addresses") or 0)
    except (TypeError, ValueError): uniq = 0
    if uniq <= 1:                           return None, "Single-building operator"
    if re.fullmatch(r"OWNER\.?", name, re.I): return None, "Generic OWNER entry"
    return row, None


def units_score(units, tier_min, tier_max):
    """Linear inside the tier (10–20 pts). Open-ended tiers ramp then cap at 20."""
    try: u = int(units or 0)
    except (TypeError, ValueError): u = 0
    if tier_max is None:                                   # 200+
        if u >= 500: return 20.0
        return round(10 + (u - tier_min) / 300 * 10, 2)
    span = tier_max - tier_min
    if span <= 0: return 10.0
    return round(max(0, min(20, ((u - tier_min) / span) * 10 + 10)), 2)


def score(row, tier_min, tier_max):
    try: units = int(row.get("Total Units") or 0)
    except (TypeError, ValueError): units = 0
    try: uniq  = int(row.get("Unique Addresses") or 0)
    except (TypeError, ValueError): uniq  = 0
    counties = (row.get("Counties") or "").strip()
    n_counties = sum(1 for c in counties.split(",") if c.strip()) if counties else 0
    phone = (row.get("Phone") or "").strip()
    name  = (row.get("Licensee Name") or "").upper()

    s_dist   = round(min(uniq / units, 1) * 40, 2) if units > 0 else 0
    s_units  = units_score(units, tier_min, tier_max)
    s_county = 12 if n_counties >= 3 else 8 if n_counties == 2 else 3 if n_counties == 1 else 0
    s_phone  = 10 if phone else 0
    s_entity = 10 if ENTITY_RE.search(name) else 0
    s_kw     = 8  if PMKW_RE.search(name)  else 0

    return round(min(100, s_dist + s_units + s_county + s_phone + s_entity + s_kw), 1)


def process_sheet(wb, sheet_name, out_subdir, tier_min, tier_max):
    if sheet_name not in wb.sheetnames:
        print(f"  skip {sheet_name!r}: not in workbook")
        return
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  skip {sheet_name!r}: empty")
        return
    headers = [str(h) if h is not None else "" for h in rows[0]]

    raw, removed = [], []
    for r in rows[1:]:
        rec = {h: ("" if v is None else v) for h, v in zip(headers, r)}
        kept, reason = clean(rec)
        if not kept:
            removed.append((rec.get("Licensee Name", ""), reason))
            continue
        raw.append(kept)

    # Dedup by lowercased name (keep highest-units row of each)
    by_name = {}
    for rec in raw:
        key = (rec.get("Licensee Name") or "").strip().lower()
        if not key: continue
        prev = by_name.get(key)
        if not prev or int(rec.get("Total Units") or 0) > int(prev.get("Total Units") or 0):
            by_name[key] = rec
    deduped = list(by_name.values())
    n_dupe = len(raw) - len(deduped)

    out_rows = []
    for rec in deduped:
        s = score(rec, tier_min, tier_max)
        out = {c: "" for c in OUT_COLS}
        for c in ["Rank","Licensee Name","Total Units","Unique Addresses","Counties","Cities Served","Phone"]:
            out[c] = rec.get(c, "")
        out["Score"] = s
        out_rows.append(out)
    out_rows.sort(key=lambda r: -float(r["Score"] or 0))

    out_dir = ROOT / out_subdir
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / "source_cleaned_scored.csv"
    with open(out_path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS)
        w.writeheader()
        w.writerows(out_rows)

    print(f"  {sheet_name:18s}  raw={len(rows)-1:>5d}  removed={len(removed):>4d}  dedup={n_dupe:>3d}  kept={len(out_rows):>4d}  →  {out_path.relative_to(ROOT)}")


def main():
    if not XLSX.exists():
        sys.exit(f"✗ Master xlsx not found at {XLSX}")
    print(f"→ Reading {XLSX.relative_to(ROOT)}")
    wb = openpyxl.load_workbook(XLSX, read_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    print()
    for sheet_name, (out_subdir, tmin, tmax) in TIERS.items():
        process_sheet(wb, sheet_name, out_subdir, tmin, tmax)
    print()
    print("✓ Done.")


if __name__ == "__main__":
    main()
